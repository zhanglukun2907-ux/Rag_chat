import os
os.environ["HF_ENDPOINT"] = "https://hf-mirror.com"

import jieba
import json
import chromadb
import gradio as gr
from openai import OpenAI
from rank_bm25 import BM25Okapi
from sentence_transformers import SentenceTransformer, CrossEncoder

# ==========================================
# 配置
# ==========================================
client = OpenAI(
    api_key="sk-uupfypofhmpnpbbabxsanbmguibjyqsepyulxrnqxmhuaecx",
    base_url="https://api.siliconflow.cn/v1"
)

# ==========================================
# 加载模型
# ==========================================
print("正在加载Embedding模型...")
embedding_model = SentenceTransformer("BAAI/bge-small-zh-v1.5")
print("Embedding模型加载完成")

print("正在加载Rerank模型...")
reranker = CrossEncoder("BAAI/bge-reranker-base")
print("Rerank模型加载完成")

# ==========================================
# ChromaDB
# ==========================================
chroma_client = chromadb.PersistentClient(path="./chroma_db")
collection = chroma_client.get_or_create_collection(name="taxi_knowledge_base")

# ==========================================
# 全局变量
# ==========================================
all_chunks = []
bm25 = None

# ==========================================
# 多格式文件读取
# ==========================================
def read_file_content(filepath):
    """根据扩展名读取不同格式文件"""
    ext = os.path.splitext(filepath)[1].lower()
    if ext in ['.txt', '.md']:
        with open(filepath, "r", encoding="utf-8") as f:
            return f.read()
    elif ext == '.docx':
        from docx import Document
        doc = Document(filepath)
        return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
    elif ext == '.pdf':
        from PyPDF2 import PdfReader
        reader = PdfReader(filepath)
        return "\n".join([page.extract_text() or "" for page in reader.pages])
    elif ext in ['.xlsx', '.xls']:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True)
        lines = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            lines.append(f"=== {sheet} ===")
            for row in ws.iter_rows(values_only=True):
                line = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if line.strip().replace("|", "").strip():
                    lines.append(line)
        wb.close()
        return "\n".join(lines)
    elif ext == '.csv':
        with open(filepath, "r", encoding="utf-8") as f:
            return f.read()
    elif ext == '.xmind':
        from xmindparser import parse
        data = parse(filepath)
        lines = []
        def walk(node, depth=0):
            title = node.get("title", "")
            if title:
                lines.append("  " * depth + title)
            for child in node.get("children", []):
                walk(child, depth + 1)
            for child in node.get("attached", []):
                walk(child, depth + 1)
        for sheet in data:
            if "topic" in sheet:
                walk(sheet["topic"])
        return "\n".join(lines)
    else:
        return ""

# ==========================================
# 增量加载跟踪
# ==========================================
TRACKER_FILE = "./knowledge/.processed.json"

def get_file_mtime(filepath):
    return os.path.getmtime(filepath)

def load_tracker():
    if os.path.exists(TRACKER_FILE):
        with open(TRACKER_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_tracker(tracker):
    os.makedirs(os.path.dirname(TRACKER_FILE), exist_ok=True)
    with open(TRACKER_FILE, "w", encoding="utf-8") as f:
        json.dump(tracker, f, ensure_ascii=False, indent=2)

# ==========================================
# 知识库加载（增量）
# ==========================================
def load_knowledge_folder(folder_path="./knowledge", chunk_size=800, chunk_overlap=150):
    global all_chunks, bm25
    all_chunks = []
    tracker = load_tracker()
    new_tracker = {}
    changed = False

    print(f"开始加载知识库: {folder_path}")
    for filename in sorted(os.listdir(folder_path)):
        if filename.startswith("."):
            continue
        filepath = os.path.join(folder_path, filename)
        if os.path.isdir(filepath):
            continue

        mtime = get_file_mtime(filepath)
        old_info = tracker.get(filename)

        # 文件没变，跳过
        if old_info and old_info.get("mtime") == mtime:
            print(f"  跳过(未变化): {filename}")
            new_tracker[filename] = old_info
            continue

        changed = True
        text = read_file_content(filepath)
        if not text.strip():
            print(f"  跳过(内容为空): {filename}")
            continue

        # 删除该文件的旧chunk
        if old_info and "chunk_ids" in old_info:
            try:
                collection.delete(ids=old_info["chunk_ids"])
                print(f"  删除旧chunk: {len(old_info['chunk_ids'])} 条")
            except:
                pass

        chunks = []
        start = 0
        while start < len(text):
            end = start + chunk_size
            chunk = text[start:end]
            if end < len(text):
                last_newline = chunk.rfind('\n')
                if last_newline > chunk_size // 2:
                    chunk = text[start:start + last_newline]
                    end = start + last_newline
            chunks.append(f"[来源: {filename}]\n{chunk.strip()}")
            start = end - chunk_overlap

        valid_chunks = [c for c in chunks if c.replace(f"[来源: {filename}]", "").strip()]
        all_chunks.extend(valid_chunks)
        new_tracker[filename] = {"mtime": mtime, "count": len(valid_chunks)}
        print(f"  {filename}: {len(valid_chunks)} 块(新/已更新)")

    # 处理被删除的文件
    for old_file in tracker:
        if old_file not in new_tracker:
            if "chunk_ids" in tracker[old_file]:
                try:
                    collection.delete(ids=tracker[old_file]["chunk_ids"])
                    print(f"  删除(文件已移除): {old_file}")
                except:
                    pass
            changed = True

    if not changed and not all_chunks:
        print("所有文件未变化，跳过向量化")
    elif all_chunks:
        print(f"新增/更新 {len(all_chunks)} 块，开始生成向量...")
        embeddings = embedding_model.encode(
            all_chunks, show_progress_bar=True, normalize_embeddings=True
        ).tolist()
        existing_ids = collection.get()["ids"]
        start_id = len(existing_ids) if existing_ids else 0
        ids = [f"chunk_{start_id + i}" for i in range(len(all_chunks))]
        collection.add(ids=ids, documents=all_chunks, embeddings=embeddings)

        # 记录chunk_ids到tracker
        for filename in new_tracker:
            count = new_tracker[filename].get("count", 0)
            if count > 0 and "chunk_ids" not in new_tracker[filename]:
                new_tracker[filename]["chunk_ids"] = ids[:count]
                ids = ids[count:]

        print("向量库写入完成")

    # BM25始终重建（轻量操作）
    all_data = collection.get()
    all_chunks = all_data["documents"]
    if all_chunks:
        tokenized_corpus = [list(jieba.cut(chunk)) for chunk in all_chunks]
        bm25 = BM25Okapi(tokenized_corpus)
    print(f"BM25索引构建完成，共 {len(all_chunks)} 块")
    save_tracker(new_tracker)
    print("知识库加载成功")

# ==========================================
# 混合检索 + Rerank
# ==========================================
def hybrid_search(query, top_k=3, candidate_k=10):
    if bm25 is None:
        raise RuntimeError("BM25尚未初始化")

    # 向量检索
    query_embedding = embedding_model.encode(
        [query], normalize_embeddings=True
    ).tolist()
    vec_res = collection.query(
        query_embeddings=query_embedding,
        n_results=min(candidate_k, len(all_chunks))
    )
    vec_docs = vec_res["documents"][0]

    # BM25检索
    token_query = list(jieba.cut(query))
    scores = bm25.get_scores(token_query)
    top_idx = sorted(
        range(len(scores)), key=lambda i: scores[i], reverse=True
    )[:candidate_k]
    bm25_docs = [all_chunks[i] for i in top_idx]

    # 合并去重
    results = []
    seen = set()
    for doc in vec_docs + bm25_docs:
        if doc not in seen:
            seen.add(doc)
            results.append(doc)

    # Rerank精排
    if len(results) == 0:
        return []
    pairs = [[query, doc] for doc in results]
    rerank_scores = reranker.predict(pairs)
    ranked = sorted(
        zip(results, rerank_scores), key=lambda x: x[1], reverse=True
    )
    return [doc for doc, score in ranked[:top_k]]
import datetime

LOG_FILE = "./knowledge/chat_log.csv"

def log_conversation(question, answer):
    """记录用户问答"""
    os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # 转义CSV中的逗号和换行
    q = question.replace('"', '""').replace('\n', ' ')
    a = answer.replace('"', '""').replace('\n', ' ')
    header_needed = not os.path.exists(LOG_FILE)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        if header_needed:
            f.write("时间,问题,回答,反馈\n")
        f.write(f'"{timestamp}","{q}","{a}",""\n')
# ==========================================
# 问答
# ==========================================
def respond(message, history):
    if not message.strip():
        yield "请输入问题"
        return
    try:
        context_list = hybrid_search(message, top_k=3)
        if not context_list:
            yield "未检索到相关内容，请换个问题试试"
            return
        context_str = "\n\n------------------\n\n".join(context_list)

        # 构建messages，带最近5轮对话历史
        messages = []
        messages.append({
            "role": "system",
            "content": f"你是一个专业文档问答助手。要求：1.只引用参考内容 2.不得补充推测 3.保持专业术语 4.完整句子回答 5.不输出无意义字符 6.无信息则说'文档中无相关信息'\n\n参考内容：\n{context_str}"
        })

        # Gradio 4.x history格式：[{"role":"user","content":"..."}, {"role":"assistant","content":"..."}]
        recent = history[-5:] if history else []
        for msg in recent:
            if isinstance(msg, dict) and msg.get("role") in ("user", "assistant"):
                messages.append({"role": msg["role"], "content": msg["content"]})

        messages.append({"role": "user", "content": message})

        stream = client.chat.completions.create(
            model="Qwen/Qwen2.5-7B-Instruct",
            messages=messages,
            temperature=0, max_tokens=1024, stream=True
        )
        result = ""
        for chunk in stream:
            if chunk.choices[0].delta.content:
                result += chunk.choices[0].delta.content
                yield result
    except Exception as e:
        yield f"错误：{str(e)}"


# ==========================================
# 主程序
# ==========================================
if __name__ == "__main__":
    os.makedirs("./knowledge", exist_ok=True)
    load_knowledge_folder(folder_path="./knowledge")
    demo = gr.ChatInterface(
        fn=respond,
        title="个人知识库问答",
        description="支持txt/md/docx/pdf/xlsx/csv/xmind，增量加载"
    )
    demo.launch(server_name="127.0.0.1", server_port=7860)
